import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import timedelta
import re, io, base64, os, json, time

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill

st.set_page_config(page_title="CSU Online Analytics Dashboard", page_icon="🎓", layout="wide")

# --- safe rerun helper ---
def safe_rerun():
    try:
        st.rerun()
    except Exception:
        try:
            st.experimental_rerun()
        except Exception:
            pass

# ========================= Branding Persistence =========================
BRAND_FILE = "branding.json"
ASSETS_DIR = "assets"
DEFAULTS = {
    "primary": "#1E4D2B",   # CSU Green
    "gold":    "#C8C372",   # CSU Gold
    "text":    "#111111",
    "bg":      "#FFFFFF",   # White
    "card":    "#F6F7F4",
    "mode":    "light",
    "logo":    f"{ASSETS_DIR}/csu_logo.png",
}

def load_branding():
    if os.path.exists(BRAND_FILE):
        try:
            with open(BRAND_FILE, "r") as f:
                data = json.load(f)
                return {**DEFAULTS, **data}
        except Exception:
            return DEFAULTS.copy()
    return DEFAULTS.copy()

def save_branding(cfg: dict, uploaded_logo=None):
    with open(BRAND_FILE, "w") as f:
        json.dump(cfg, f, indent=2)
    if uploaded_logo is not None:
        os.makedirs(ASSETS_DIR, exist_ok=True)
        logo_path = os.path.join(ASSETS_DIR, "csu_logo.png")
        with open(logo_path, "wb") as out:
            out.write(uploaded_logo.getbuffer())
        cfg["logo"] = logo_path
        with open(BRAND_FILE, "w") as f:
            json.dump(cfg, f, indent=2)
    return cfg

def load_logo_b64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    except Exception:
        return None

def inject_brand_css(brand, logo_b64=None):
    primary, gold, text, bg, card, mode = (
        brand["primary"], brand["gold"], brand["text"], brand["bg"], brand["card"], brand["mode"]
    )
    if mode == "dark":
        bg   = "#0E1117"
        card = "#161B22"
        text = "#E6EDF3"
    logo_html = (
        f'<img src="data:image/png;base64,{logo_b64}" style="height:48px;margin-right:10px;vertical-align:middle;" />'
        if logo_b64
        else ""
    )
    st.markdown(
        f"""
        <style>
        :root {{
          --brand-primary: {primary};
          --brand-gold: {gold};
          --brand-text: {text};
          --brand-bg: {bg};
          --brand-card: {card};
        }}
        .stApp {{ background: var(--brand-bg); color: var(--brand-text); }}
        .kpi-card {{
          background: var(--brand-card);
          border-radius: 14px;
          padding: 14px;
          box-shadow: 0 4px 18px rgba(0,0,0,.06);
          border: 1px solid rgba(0,0,0,.05);
        }}
        .stTabs [data-baseweb="tab-list"] {{ gap: 8px; }}
        .stTabs [data-baseweb="tab"] {{
          padding: 10px 16px;
          background: #fff;
          border-radius: 10px;
          border: 1px solid rgba(0,0,0,.08);
        }}
        .stTabs [aria-selected="true"] {{ background: #eef5ee; border-color: var(--brand-primary); }}

        /* Animated headline */
        .welcome-wrap {{ display:flex; align-items:center; justify-content:center; height:70vh; text-align:center; }}
        .welcome {{
            font-weight:800; font-size: clamp(28px, 6vw, 56px); line-height:1.1;
            background: linear-gradient(90deg, var(--brand-primary), var(--brand-gold), var(--brand-primary));
            -webkit-background-clip: text; -webkit-text-fill-color: transparent;
            background-size: 200% 200%; animation: shimmer 3s ease-in-out infinite;
        }}
        @keyframes shimmer {{
            0% {{background-position:0% 50%;}}
            50% {{background-position:100% 50%;}}
            100% {{background-position:0% 50%;}}
        }}
        .prompt {{ font-size: clamp(18px, 3vw, 26px); color: var(--brand-text); }}
        .ghost-btn {{
            border:2px solid var(--brand-primary);
            color:var(--brand-primary);
            padding:12px 18px; border-radius:999px; font-weight:700;
            background:white; cursor:pointer;
        }}
        .ghost-btn:hover {{ background: rgba(30,77,43,0.06); }}
        .progress {{ width:100%; height:10px; background:#e5e7eb; border-radius:999px; overflow:hidden; }}
        .bar {{ height:100%; background: linear-gradient(90deg, var(--brand-primary), var(--brand-gold)); }}
        .steptext {{ font-size:14px; color:#4b5563; margin-bottom:8px; }}
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;">{logo_html}<span style="font-weight:800;font-size:1.25rem;">CSU Online Analytics Dashboard</span></div>',
        unsafe_allow_html=True,
    )

def apply_colorway(fig):
    fig.update_layout(colorway=[st.session_state.brand["primary"], st.session_state.brand["gold"], "#2E7D32", "#8C7E3E", "#0B5D1E"])
    return fig

# Initialize branding
if "brand" not in st.session_state:
    st.session_state.brand = load_branding()
logo_b64 = load_logo_b64(st.session_state.brand["logo"]) if st.session_state.brand.get("logo") else None
inject_brand_css(st.session_state.brand, logo_b64)

# ========================= Shared helpers =========================
def kpi_card(title, value):
    st.markdown(
        f"""<div class="kpi-card">
               <div style="font-size:.9rem;color:#4b5563;margin-bottom:4px;">{title}</div>
               <div style="font-size:1.6rem;font-weight:700;">{value}</div>
            </div>""",
        unsafe_allow_html=True,
    )

def fig_layout(fig, h=420):
    fig.update_layout(
        height=h,
        margin=dict(l=30, r=10, t=40, b=35),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    return apply_colorway(fig)

def time_to_seconds(ts: str) -> int:
    if pd.isna(ts) or ts == "":
        return 0
    parts = list(map(int, str(ts).split(":")))
    while len(parts) < 3:
        parts.insert(0, 0)
    h, m, s = parts
    return h * 3600 + m * 60 + s

def seconds_to_hms(sec: float) -> str:
    return "" if pd.isna(sec) else str(timedelta(seconds=int(sec)))

def natural_key(s: str):
    return [int(ch) if ch.isdigit() else ch.lower() for ch in re.split(r"(\d+)", str(s))]

# ========================= Echo Script Parity =========================
def echo_analyze(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Duration_sec"] = df["Duration"].apply(time_to_seconds)
    df["TotalViewTime_sec"] = df["Total View Time"].apply(time_to_seconds)
    df["AvgViewTime_sec"] = df["Average View Time"].apply(time_to_seconds)
    df["Row View %"] = df["TotalViewTime_sec"] / df["Duration_sec"].replace(0, np.nan)

    grp = df.groupby("Media Name", sort=False)
    titles = list(grp.groups.keys())
    summary = pd.DataFrame(
        {
            "Media Title": titles,
            "Video Duration": [grp.get_group(t)["Duration_sec"].iloc[0] for t in titles],
            "Number of Unique Viewers": grp["User Name"].nunique().values,
            "Average View %": grp["Row View %"].mean().fillna(0).values,
            "Total View %": (grp["TotalViewTime_sec"].sum() / grp["Duration_sec"].sum()).values,
            "Total View Time": grp["TotalViewTime_sec"].sum().values,
            "Average View Time": grp["AvgViewTime_sec"].mean().values,
            "Average Total View Time": grp["TotalViewTime_sec"].mean().values,
        }
    )
    summary["sort_key"] = summary["Media Title"].apply(natural_key)
    summary = summary.sort_values("sort_key").drop(columns="sort_key").reset_index(drop=True)

    means = summary[["Video Duration", "Total View Time", "Average View Time", "Average Total View Time"]].mean()
    viewers_mean = summary["Number of Unique Viewers"].mean()
    summary.loc[len(summary)] = {
        "Media Title": "Grand Total",
        "Video Duration": means["Video Duration"],
        "Number of Unique Viewers": viewers_mean,
        "Average View %": summary["Average View %"].mean(),
        "Total View %": summary["Total View %"].mean(),
        "Total View Time": means["Total View Time"],
        "Average View Time": means["Average View Time"],
        "Average Total View Time": means["Average Total View Time"],
    }

    n = len(summary) - 1
    means2 = summary.loc[: n - 1, ["Video Duration", "Total View Time", "Average View Time", "Average Total View Time"]].mean()
    summary.loc[len(summary)] = {
        "Media Title": "Average Video Length and Watch Time",
        "Video Duration": means2["Video Duration"],
        "Number of Unique Viewers": "",
        "Average View %": summary.loc[: n - 1, "Average View %"].mean(),
        "Total View %": summary.loc[: n - 1, "Total View %"].mean(),
        "Total View Time": means2["Total View Time"],
        "Average View Time": means2["Average View Time"],
        "Average Total View Time": means2["Average Total View Time"],
    }
    return summary

def echo_build_workbook(summary_df: pd.DataFrame) -> bytes:
    # Accessibility: use default Excel conditional formatting colors
    wb = Workbook()
    ws = wb.active
    ws.title = "Media Summary"

    tmp = summary_df.copy()
    for col in ["Video Duration", "Total View Time", "Average View Time", "Average Total View Time"]:
        tmp[col] = tmp[col].apply(seconds_to_hms)
    for row in dataframe_to_rows(tmp, index=False, header=True):
        ws.append(row)

    last_row = ws.max_row
    media_count = len(summary_df) - 2

    # Duration to time format
    for r in range(2, last_row + 1):
        cell = ws[f"B{r}"]
        secs = time_to_seconds(cell.value)
        cell.value = secs / 86400.0
        cell.number_format = "hh:mm:ss"

    # Percent/time formatting
    for r in range(2, last_row + 1):
        for col in ("D", "E"):
            c = ws[f"{col}{r}"]
            if isinstance(c.value, (int, float)):
                c.number_format = "0.00%"
        for col in ("F", "G", "H"):
            ws[f"{col}{r}"].number_format = "hh:mm:ss"

    # Data bars (default color for accessibility)
    if media_count >= 1:
        bar = DataBarRule(start_type="min", end_type="max")
        ws.conditional_formatting.add(f"B2:B{1 + media_count}", bar)
        ws.conditional_formatting.add(f"D2:D{1 + media_count}", bar)

    # Charts
    chart1 = LineChart()
    chart1.title = "View % Over Time"
    chart1.style = 9
    chart1.y_axis.number_format = "0.00%"
    data1 = Reference(ws, min_col=4, min_row=1, max_row=1 + media_count)
    chart1.add_data(data1, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + media_count)
    chart1.set_categories(cats)
    ws.add_chart(chart1, "J2")

    chart2 = LineChart()
    chart2.title = "Unique Viewers Over Time"
    chart2.style = 9
    data2 = Reference(ws, min_col=3, min_row=1, max_row=1 + media_count)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    ws.add_chart(chart2, "J20")

    tbl = Table(displayName="MediaStats", ref=f"A1:H{last_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ========================= Gradebook Script Parity =========================
def gradebook_process(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    mask = df.iloc[:, 0].astype(str).str.contains("Student, Test", na=False)
    df = df[~mask].reset_index(drop=True)
    to_drop = ["Student", "ID", "SIS User ID", "SIS Login ID", "Current Grade", "Unposted Current Grade", "Unposted Final Grade"]
    df.drop(columns=[c for c in to_drop if c in df.columns], inplace=True, errors="ignore")

    drop_cols = []
    for col in df.columns:
        if col == "Final Grade":
            continue
        s = pd.to_numeric(df[col].iloc[2:], errors="coerce")
        if s.fillna(0).eq(0).all():
            drop_cols.append(col)
    df.drop(columns=drop_cols, inplace=True, errors="ignore")
    return df

def gradebook_build_workbook(df: pd.DataFrame) -> bytes:
    # Accessibility: keep default Excel conditional formatting colors
    wb = Workbook()
    ws = wb.active
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    final_grade_idx_pre = None
    for ci in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=ci).value == "Final Grade":
            final_grade_idx_pre = ci
            break

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.column == final_grade_idx_pre:
                continue
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ""):
                cell.value = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.column == final_grade_idx_pre:
                continue
            if isinstance(cell.value, str) and cell.value.strip():
                txt = cell.value.replace(",", "")
                try:
                    cell.value = float(txt)
                except ValueError:
                    pass

    data_last_row = ws.max_row
    for ci in range(1, ws.max_column + 1):
        if ci == final_grade_idx_pre:
            continue
        hdr = ws.cell(row=2, column=ci)
        if isinstance(hdr.value, str) and "(read only)" in hdr.value:
            nums = [
                ws.cell(row=r, column=ci).value
                for r in range(3, data_last_row + 1)
                if isinstance(ws.cell(row=r, column=ci).value, (int, float))
            ]
            if nums:
                hdr.value = max(nums)

    ws.insert_cols(1)
    ws["A1"] = "Row Titles"

    final_grade_idx = None
    for ci in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=ci).value == "Final Grade":
            final_grade_idx = ci
            break

    ws["A2"] = "Points Possible"
    original_last_data_row = ws.max_row
    avg_row = original_last_data_row + 1
    avg0_row = original_last_data_row + 2
    ws[f"A{avg_row}"] = "Average"
    ws[f"A{avg0_row}"] = "Average Excluding Zeros"

    max_col = ws.max_column
    for col in range(2, max_col + 1):
        if col == final_grade_idx:
            continue
        letter = get_column_letter(col)
        data_rng = f"{letter}3:{letter}{original_last_data_row}"
        header = f"{letter}$2"
        c_avg = ws[f"{letter}{avg_row}"]
        c_avg.value = f"=AVERAGE({data_rng})/{header}"
        c_avg.number_format = "0.00%"
        c_avg0 = ws[f"{letter}{avg0_row}"]
        c_avg0.value = f'=AVERAGEIF({data_rng},">0")/{header}'
        c_avg0.number_format = "0.00%"

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row in (avg_row, avg0_row):
        rng = f"B{row}:{get_column_letter(max_col)}{row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0.9"], fill=green))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["0.8", "0.9"], fill=yellow))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0.8"], fill=red))

    count_row = avg0_row + 1
    pct_row = avg0_row + 2
    ws[f"A{count_row}"] = "Count of F"
    ws[f"A{pct_row}"] = "Percent of F"
    fg_letter = get_column_letter(final_grade_idx)
    ws.cell(row=count_row, column=final_grade_idx).value = f'=COUNTIF({fg_letter}3:{fg_letter}{original_last_data_row},"F")'
    total_students = original_last_data_row - 2
    ws.cell(row=pct_row, column=final_grade_idx).value = f"={fg_letter}{count_row}/{total_students}"
    ws.cell(row=pct_row, column=final_grade_idx).number_format = "0.00%"

    table_end = get_column_letter(max_col) + str(original_last_data_row)
    tbl = Table(displayName="GradesTable", ref=f"A1:{table_end}")
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ========================= Gradebook Dashboard Prep =========================
def parse_missing_excused(series: pd.Series):
    s = series.astype(str).str.strip()
    excused = s.str.upper().isin(["EX", "EXCUSED"])
    blanks = s.isin(["", "nan", "NaN", "-", "—", "–"])
    nums = pd.to_numeric(s.str.replace("%", "", regex=False), errors="coerce")
    missing = blanks | nums.isna()
    return nums, missing, excused

def extract_category(c):
    base = re.sub(r"\(\d+\)$", "", c).strip()
    if ":" in base:
        return base.split(":")[0].strip()
    if "-" in base:
        return base.split("-")[0].strip()
    return base.split()[0] if base.split() else base

def gradebook_prepare_for_dashboard(df_proc: pd.DataFrame, student_col="Student", section_col="Section", final_col="Final Grade"):
    if df_proc.shape[0] < 3:
        return None
    points = df_proc.iloc[1].copy()
    data = df_proc.iloc[2:].copy()
    if student_col in df_proc.columns:
        data[student_col] = df_proc.iloc[2:][student_col].values
    if section_col in df_proc.columns:
        data[section_col] = df_proc.iloc[2:][section_col].values

    assign_cols = [c for c in df_proc.columns if c not in [student_col, section_col, final_col]]
    num_df = pd.DataFrame(index=data.index, columns=assign_cols, dtype=float)
    miss_df = pd.DataFrame(False, index=data.index, columns=assign_cols)
    exc_df = pd.DataFrame(False, index=data.index, columns=assign_cols)
    for c in assign_cols:
        nums, miss, exc = parse_missing_excused(data[c])
        num_df[c] = nums
        miss_df[c] = miss | nums.isna()
        exc_df[c] = exc

    pts = pd.to_numeric(points[assign_cols], errors="coerce").replace(0, np.nan)
    pct_df = (num_df / pts) * 100.0

    # Align indexes to avoid boolean indexer mismatch downstream
    pct_df = pct_df.reset_index(drop=True)
    miss_df = miss_df.reset_index(drop=True)
    exc_df = exc_df.reset_index(drop=True)
    data = data.reset_index(drop=True)

    return {
        "data": data,
        "final_col": final_col if final_col in df_proc.columns else None,
        "student_col": student_col if student_col in df_proc.columns else None,
        "section_col": section_col if section_col in df_proc.columns else None,
        "assign_cols": assign_cols,
        "pct_df": pct_df,
        "missing": miss_df,
        "excused": exc_df,
        "points": pts,
        "categories": {c: extract_category(c) for c in assign_cols},
    }

# ========================= Onboarding with Progress =========================
if "stage" not in st.session_state:
    st.session_state.stage = "welcome"
if "gb_file" not in st.session_state:
    st.session_state.gb_file = None
if "echo_file" not in st.session_state:
    st.session_state.echo_file = None
if "welcome_started" not in st.session_state:
    st.session_state.welcome_started = None

def progress_ui(step: int):
    steps = {1: "Welcome", 2: "Canvas Gradebook Upload", 3: "Echo Upload"}
    pct = {1: 0.33, 2: 0.66, 3: 1.00}[step]
    st.markdown(f'<div class="steptext">Step {step}/3 — {steps[step]}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="progress"><div class="bar" style="width:{int(pct*100)}%"></div></div>', unsafe_allow_html=True)

def show_welcome():
    progress_ui(1)
    placeholder = st.empty()
    with placeholder.container():
        st.markdown(
            '<div class="welcome-wrap"><div class="welcome">Welcome to the CSU Online Analytics Dashboard!</div></div>',
            unsafe_allow_html=True,
        )
    clicked = st.button("Continue ▶️", type="primary", use_container_width=True)
    if clicked:
        st.session_state.stage = "gradebook"
        safe_rerun()
    if st.session_state.welcome_started is None:
        st.session_state.welcome_started = time.time()
    elif time.time() - st.session_state.welcome_started >= 5:
        st.session_state.stage = "gradebook"
        safe_rerun()

def show_gradebook_prompt():
    progress_ui(2)
    st.markdown(
        '<div class="welcome-wrap"><div class="prompt">Please Upload Your <b>Canvas Gradebook</b> Data</div></div>',
        unsafe_allow_html=True,
    )
    gb = st.file_uploader("Drag & drop your Gradebook CSV here", type=["csv"], key="gb_onboard")
    if gb is not None:
        st.session_state.gb_file = gb.getvalue()
        st.session_state.stage = "echo"
        safe_rerun()

def show_echo_prompt():
    progress_ui(3)
    st.markdown(
        '<div class="welcome-wrap"><div class="prompt">Please Upload Your <b>Echo</b> Data</div></div>',
        unsafe_allow_html=True,
    )
    ec = st.file_uploader("Drag & drop your Echo CSV here", type=["csv"], key="echo_onboard")
    if ec is not None:
        st.session_state.echo_file = ec.getvalue()
        st.session_state.stage = "dashboard"
        safe_rerun()

# ========================= Render by stage =========================
if st.session_state.stage == "welcome":
    show_welcome()

elif st.session_state.stage == "gradebook":
    show_gradebook_prompt()

elif st.session_state.stage == "echo":
    show_echo_prompt()

else:
    # DASHBOARD
    st.success("Files loaded. Welcome to your dashboard. Use the tabs below.")
    tab_gb, tab_echo, tab_profile = st.tabs(["📘 Gradebook", "🎬 Echo", "👤 Student Profile"])

    # ===== Gradebook Tab =====
    with tab_gb:
        st.subheader("Gradebook")
        if st.session_state.gb_file is None:
            st.info("No gradebook file found. Restart to upload.")
        else:
            from io import StringIO
            df_raw = pd.read_csv(StringIO(st.session_state.gb_file.decode("utf-8")))

            # mapping optional; raw should already have "Final Grade" if present
            df_proc = gradebook_process(df_raw)
            st.dataframe(df_proc.head(20), use_container_width=True)

            try:
                xbytes = gradebook_build_workbook(df_proc)
                st.download_button(
                    "⬇️ Download Excel (script formulas & formatting)",
                    data=xbytes,
                    file_name="Gradebook_Analyzed.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"Excel build failed: {e}")

            prep = gradebook_prepare_for_dashboard(df_proc)
            if prep:
                data = prep["data"]
                final_c = prep["final_col"]
                student_c = prep["student_col"]
                section_c = prep["section_col"]
                assign_cols = prep["assign_cols"]
                pct_df = prep["pct_df"]
                missing = prep["missing"]
                excused = prep["excused"]
                categories = prep["categories"]

                if "gb_filters" not in st.session_state:
                    st.session_state.gb_filters = {"section": "All", "students": [], "categories": []}

                fc1, fc2, fc3 = st.columns([1.2, 1.2, 2])
                with fc1:
                    if section_c and section_c in data.columns:
                        opts = ["All"] + sorted([x for x in data[section_c].dropna().unique().tolist() if x != ""])
                        sel = st.session_state.gb_filters.get("section", "All")
                        idx = opts.index(sel) if sel in opts else 0
                        st.session_state.gb_filters["section"] = st.selectbox("Section", opts, index=idx)
                with fc2:
                    cat_list = sorted(set(categories.values()))
                    st.session_state.gb_filters["categories"] = st.multiselect(
                        "Categories", options=cat_list, default=st.session_state.gb_filters["categories"]
                    )
                with fc3:
                    if student_c and student_c in data.columns:
                        options = data[student_c].dropna().astype(str).tolist()
                        prev = st.session_state.gb_filters.get("students", [])
                        default = [s for s in prev if s in options][:8]
                        st.session_state.gb_filters["students"] = st.multiselect(
                            "Students (focus)", options=options, default=default, max_selections=8
                        )

                # -------- Robust mask & slicing (no .loc with foreign index) --------
                rows_mask = pd.Series(True, index=data.index)
                if st.session_state.gb_filters["section"] != "All" and section_c and section_c in data.columns:
                    rows_mask &= (data[section_c] == st.session_state.gb_filters["section"])

                # Build aligned boolean arrays per frame, then use .iloc
                mask_data = rows_mask.reindex(data.index,   fill_value=False).to_numpy(dtype=bool)
                mask_pct  = rows_mask.reindex(pct_df.index, fill_value=False).to_numpy(dtype=bool)
                mask_mis  = rows_mask.reindex(missing.index, fill_value=False).to_numpy(dtype=bool)
                mask_exc  = rows_mask.reindex(excused.index, fill_value=False).to_numpy(dtype=bool)

                data_f = data.iloc[mask_data].reset_index(drop=True)
                pct_f  = pct_df.iloc[mask_pct]
                miss_f = missing.iloc[mask_mis]
                exc_f  = excused.iloc[mask_exc]
                # -------------------------------------------------------------------

                assn_f = assign_cols[:]
                if st.session_state.gb_filters["categories"]:
                    assn_f = [a for a in assign_cols if categories[a] in st.session_state.gb_filters["categories"]]
                if len(assn_f) == 0:
                    assn_f = assign_cols[:]

                k1, k2, k3, k4 = st.columns(4)
                kpi_card("Students", data_f.shape[0])
                kpi_card("Assignments", len(assn_f))
                if final_c and final_c in data_f.columns:
                    vals = data_f[final_c].astype(str)
                    kpi_card("Unique Letter Grades", vals.nunique())
                    kpi_card("Count of F", int((vals == "F").sum()))
                else:
                    kpi_card("Unique Letter Grades", "—")
                    kpi_card("Count of F", "—")

                incl = pct_f[assn_f].fillna(0.0)
                excl = pct_f[assn_f].copy()
                avg_incl = incl.mean().sort_values()
                avg_excl = excl.mean().reindex(avg_incl.index)
                labels = [re.sub(r"\(\d+\)$", "", a) for a in avg_incl.index]
                fig = go.Figure()
                fig.add_bar(x=labels, y=avg_incl.values, name="Including Missing")
                fig.add_bar(x=labels, y=avg_excl.values, name="Excluding Missing")
                fig.update_xaxes(tickangle=45)
                st.plotly_chart(fig_layout(fig, h=420), use_container_width=True)

                st.subheader("Missing / Excused Heatmap")
                if not miss_f.empty and student_c and student_c in data_f.columns:
                    mat = miss_f[assn_f].astype(int).values.astype(float) - 0.5 * exc_f[assn_f].astype(int).values
                    fig_hm = px.imshow(
                        mat,
                        labels=dict(x="Assignments", y="Students", color="Status"),
                        x=[re.sub(r"\(\d+\)$", "", a)[:18] for a in assn_f],
                        y=data_f[student_c].tolist(),
                        aspect="auto",
                    )
                    st.plotly_chart(fig_layout(fig_hm, h=min(800, 40 + 22 * data_f.shape[0])), use_container_width=True)

                st.subheader("Student Trajectories")
                if student_c and student_c in data_f.columns:
                    picks = st.session_state.gb_filters["students"]
                    if picks:
                        figT = go.Figure()
                        for s in picks:
                            row = pct_f.loc[data_f[student_c] == s, assn_f]
                            if not row.empty:
                                figT.add_scatter(
                                    x=[re.sub(r"\(\d+\)$", "", a) for a in assn_f],
                                    y=row.iloc[0].values,
                                    mode="lines+markers",
                                    name=s,
                                )
                        figT.update_yaxes(title_text="Score (%)", range=[0, 100])
                        figT.update_xaxes(tickangle=45, title_text="Assignments")
                        st.plotly_chart(fig_layout(figT, h=420), use_container_width=True)
                    else:
                        st.caption("Select students above to view trajectories.")

                st.subheader("Assignment Correlations (completed work only)")
                if len(assn_f) >= 2:
                    corr = pct_f[assn_f].replace(0, np.nan).corr()
                    st.plotly_chart(
                        fig_layout(px.imshow(corr, text_auto=".2f", aspect="auto", color_continuous_scale="Greens"), h=520),
                        use_container_width=True,
                    )

    # ===== Echo Tab =====
    with tab_echo:
        st.subheader("Echo")
        if st.session_state.echo_file is None:
            st.info("No Echo file found. Restart to upload.")
        else:
            from io import StringIO
            df = pd.read_csv(StringIO(st.session_state.echo_file.decode("utf-8")), dtype=str)
            required = ["Media Name", "Duration", "User Name", "Total View Time", "Average View Time"]
            missing_cols = [c for c in required if c not in df.columns]
            if missing_cols:
                st.error(f"Missing required columns: {missing_cols}")
            else:
                summary = echo_analyze(df)
                try:
                    xbytes = echo_build_workbook(summary)
                    st.download_button(
                        "⬇️ Download Excel (script formatting & charts)",
                        data=xbytes,
                        file_name="Echo_Analyzed.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception as e:
                    st.error(f"Excel build failed: {e}")

                disp = summary.copy()
                for col in ["Video Duration", "Total View Time", "Average View Time", "Average Total View Time"]:
                    disp[col] = disp[col].apply(
                        lambda x: seconds_to_hms(x) if isinstance(x, (int, float, np.integer, np.floating)) else x
                    )
                st.dataframe(disp, use_container_width=True)

                media_count = max(0, len(summary) - 2)
                if media_count > 0:
                    main = summary.iloc[:media_count].copy()
                    fig1 = go.Figure()
                    fig1.add_trace(
                        go.Scatter(x=main["Media Title"], y=main["Average View %"], mode="lines+markers", name="Average View %")
                    )
                    fig1.update_yaxes(tickformat=".0%")
                    st.plotly_chart(fig_layout(fig1, h=420), use_container_width=True)

                    fig2 = go.Figure()
                    fig2.add_trace(
                        go.Scatter(x=main["Media Title"], y=main["Number of Unique Viewers"], mode="lines+markers", name="Unique Viewers")
                    )
                    st.plotly_chart(fig_layout(fig2, h=420), use_container_width=True)

    # ===== Student Profile Tab =====
    with tab_profile:
        st.subheader("Student Profile")
        st.caption("Select students in the Gradebook tab to view their trajectories and use KPIs above.")
